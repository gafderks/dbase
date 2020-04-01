from datetime import datetime

from django.http import HttpResponse
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from booking.models import PartOfDay
from booking.views import EventListView


class EventExcelView(EventListView):
    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response[
            "Content-Disposition"
        ] = "attachment; filename={bookings} {event} ({group}) - {downloaded} {date}.xlsx".format(
            bookings=_("Material bookings"),
            event=context["current_event"].name,
            group=context["current_group"].name
            if context["current_group"]
            else _("All groups"),
            downloaded=_("downloaded on"),
            date=datetime.now().strftime("%Y-%m-%d"),
        )

        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = _("Bookings")

        columns = [
            (_("Day"), 10),
            (_("Daypart"), 10),
            (_("Amount"), 8),
            (_("Material"), 30),
            (_("Material category"), 25),
            (_("GM"), 10),
            (_("Game"), 30),
            (_("Group"), 12),
            (_("Workweek"), 10),
            (_("List"), 25),
            (_("Comment"), 50),
        ]

        row_num = 1
        last_cell = None

        wrapped_alignment = Alignment(vertical="top", wrap_text=True)
        left_alignment = Alignment(horizontal="left")
        right_alignment = Alignment(horizontal="right")
        centered_alignment = Alignment(horizontal="center")

        # Assign the titles for each cell of the header
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            # Set column width
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width

            last_cell = cell

        for day, part_of_days in context["list_views"].items():
            for part_of_day, filters in part_of_days.items():
                for filter in filters:
                    for booking in filter.bookings:
                        # Skip the header row
                        row_num += 1

                        row = [
                            day,  # 1
                            str(PartOfDay.name_from_code(part_of_day)),  # 2
                            float(booking.amount)
                            if booking.amount.isnumeric()
                            else booking.amount,  # 3
                            str(booking),  # 4
                            str(booking.display_category),  # 5
                            int(
                                booking.material.gm if booking.material else False
                            ),  # 6
                            booking.game.name,  # 7
                            booking.game.group.name,  # 8
                            int(booking.workweek),  # 9
                            str(filter.name),  # 10
                            booking.comment,  # 11
                        ]

                        # Assign the data for each cell of the row
                        for col_num, cell_value in enumerate(row, 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.value = cell_value
                            cell.alignment = wrapped_alignment
                            if col_num == 1:
                                cell.number_format = "dddd"
                                cell.alignment = left_alignment
                            if col_num == 3:
                                cell.alignment = right_alignment
                            if col_num == 6 or col_num == 9:
                                cell.number_format = '"{}";;"{}";'.format(
                                    _("Yes"), _("No")
                                )
                                cell.alignment = centered_alignment

                            last_cell = cell

        tab = Table(displayName="Table1", ref="A1:{}".format(last_cell.coordinate))

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleMedium15",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

        workbook.save(response)

        return response
